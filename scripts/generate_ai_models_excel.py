"""
Generate simplified AI Models reference Excel workbook.
Run: python scripts/generate_ai_models_excel.py
Output: docs/AI_MODELS_AND_MODULES_REFERENCE.xlsx
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "docs" / "AI_MODELS_REFERENCE.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# Model | Used For
MODELS = [
    # Video / Classroom pipeline
    ("OpenCV + ffmpeg", "Extract frames from uploaded classroom video"),
    ("YOLOv8 / YOLOv11", "Detect teacher and children in classroom video"),
    ("MediaPipe Face Detection", "Count children and improve attendance accuracy"),
    ("MediaPipe Pose", "Estimate child attention and body posture"),
    ("ByteTrack / BoT-SORT", "Track same child across frames — avoid double counting"),
    ("SlowFast / X3D", "Detect participation actions (listening, hand raise, distracted)"),
    ("CLIP / SigLIP", "Identify activity type — storytelling, rhyme, drawing, etc."),
    ("Whisper", "Convert classroom audio to Telugu/English transcript"),
    ("wav2vec2-XLS-R", "Analyze speech quality and language in low-resource Telugu"),
    ("HuBERT / Emotion model", "Measure teacher voice tone and storytelling expression"),
    ("pyannote.audio", "Separate teacher speech from child responses in audio"),
    ("WebRTC VAD", "Detect speech vs silence before audio analysis"),
    ("sentence-transformers", "Match session content with planned syllabus/activity"),
    ("Llama 3 / Mistral (LLM)", "Generate teacher/child observations and coaching tips"),
    ("GPT-4o-mini / Gemini (LLM)", "Create supervisor summary and support guidance"),
    ("DeepFace / FER+ (optional)", "Optional facial expression signal for engagement"),
    # Current app (demo — no real ML API)
    ("Demo heuristic engine (current app)", "Hackathon demo scores when real models are not connected"),
    ("demoStorytellingTemplate", "Fixed storytelling demo output for jury presentation"),
    # Public portal
    ("IndicBERT / sentiment model (production)", "Analyze citizen Share Experience feedback sentiment"),
    ("IndicTrans (production)", "Translate Telugu/English citizen feedback text"),
    # Attendance
    ("YOLO + tracker (attendance)", "Suggest preschool child count from session video for attendance"),
]


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "AI Models"

    headers = ["Model", "Used For"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN

    for model, used_for in MODELS:
        ws.append([model, used_for])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = WRAP
            cell.border = THIN

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 62
    ws.freeze_panes = "A2"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Created: {OUT}")


if __name__ == "__main__":
    main()
